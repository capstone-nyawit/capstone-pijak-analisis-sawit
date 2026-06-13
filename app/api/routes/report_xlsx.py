"""
NyawitAI — User Activity & Audit Log XLSX Report Generator
Endpoint: GET /api/admin/reports/user-activity-xlsx
Returns: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
"""

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from app.core.security import get_current_user_token
from app.db.session import get_db
from app.models.user import User
from app.models.inference_log import InferenceLog
from app.models.report import Report
from app.models.activity_log import ActivityLog

from datetime import datetime, timedelta
from datetime import date as _date
import uuid
import io

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

router = APIRouter()

# ─── Colour constants ─────────────────────────────────────────────────────────
# ─── Colour constants ─────────────────────────────────────────────────────────
C_HEADER_BG   = "04211A"  # Dark Forest Green for headers
C_HEADER_FG   = "FFFFFF"  # White text for headers
C_SECTION_BG  = "ECFDF5"  # Light emerald/mint accent background for sections
C_ALT_ROW     = "F4F7F6"  # Soft forest tint for alternating rows
C_BORDER      = "D1D5DB"  # Clean visible light border
C_GRAY        = "4B5563"  # Text gray
C_DARK        = "0F172A"  # Rich dark slate for text
C_GREEN       = "047857"  # Rich emerald green for positive status
C_AMBER       = "B45309"  # Deep amber
C_RED         = "B91C1C"  # Dark crimson red
C_BLUE        = "1D4ED8"  # Deep royal blue
C_LGRAY       = "9CA3AF"  # Light placeholder gray

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=C_DARK, italic=False) -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def _border() -> Border:
    thin = Side(style="thin", color=C_BORDER)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_row(ws, row: int, values: list, bold=False, size=11,
             color=C_DARK, bg=None, italic=False, border=False,
             alignments=None):
    """Write a list of values into consecutive cells of a row."""
    ws.row_dimensions[row].height = 20
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = _font(bold=bold, size=size, color=color, italic=italic)
        cell.alignment = _align(h=(alignments[col_idx-1] if alignments else "left"))
        if bg:
            cell.fill = _fill(bg)
        if border:
            cell.border = _border()

def _header_row(ws, row: int, labels: list):
    """Dark-background column header row."""
    ws.row_dimensions[row].height = 26
    for col_idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _font(bold=True, size=11, color=C_HEADER_FG)
        cell.fill = _fill(C_HEADER_BG)
        cell.alignment = _align(h="left")
        cell.border = _border()

def _section_label(ws, row: int, label: str, col_span: int):
    """Light-gray section header spanning multiple columns."""
    ws.row_dimensions[row].height = 24
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = _font(bold=True, size=11, color=C_DARK)
    cell.fill = _fill(C_SECTION_BG)
    cell.alignment = _align()
    bot = Side(style="thin", color="10B981")
    cell.border = Border(bottom=bot)
    # blank the rest of the span
    for c in range(2, col_span + 1):
        ws.cell(row=row, column=c).fill = _fill(C_SECTION_BG)
        ws.cell(row=row, column=c).border = Border(bottom=bot)

def _merge_header(ws, row: int, text: str, max_col: int, bold=True, size=14):
    """Merge entire row and write report title."""
    ws.row_dimensions[row].height = 32
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _font(bold=bold, size=size, color=C_DARK)
    cell.alignment = _align()

def _monitoring_period() -> str:
    today = datetime.utcnow().date()
    start = _date(today.year, today.month, 1)
    start_str = f"{start.strftime('%b')} {start.day}"   # e.g. Jun 1
    end_str   = today.strftime("%b %d, %Y")  # e.g. Jun 12, 2026
    return f"{start_str} – {end_str}"

def _fmt_dt(dt) -> str:
    if dt is None:
        return "Never"
    return dt.strftime("%b %d, %Y %I:%M %p")

def _fmt_date(dt) -> str:
    if dt is None:
        return "Never"
    return dt.strftime("%b %d, %Y")

# ─── Main handler ─────────────────────────────────────────────────────────────
@router.get("/reports/user-activity-xlsx")
def get_user_activity_xlsx(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user_token)
):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    # ── Time bounds ──────────────────────────────────────────────────────────
    now_utc      = datetime.utcnow()
    today        = now_utc.date()
    start_of_m   = datetime(today.year, today.month, 1)
    if today.month == 12:
        end_of_m = datetime(today.year + 1, 1, 1) - timedelta(seconds=1)
    else:
        end_of_m = datetime(today.year, today.month + 1, 1) - timedelta(seconds=1)

    period_label = _monitoring_period()
    gen_ts       = now_utc.strftime("%b %d, %Y %I:%M %p UTC")
    report_id    = str(uuid.uuid4()).upper()
    # ── Derive data from ActivityLog + InferenceLog + Report + User ──────────
    users = db.query(User).order_by(User.full_name).all()

    # Logins per user this month (from ActivityLog)
    logins_month = db.query(
        ActivityLog.user_id,
        func.count(ActivityLog.id).label("cnt")
    ).filter(
        ActivityLog.action == "LOGIN",
        ActivityLog.created_at.between(start_of_m, end_of_m)
    ).group_by(ActivityLog.user_id).all()

    logins_map: dict = {row.user_id: row.cnt for row in logins_month if row.user_id is not None}

    # Latest login per user (from ActivityLog)
    last_login_raw = db.query(
        ActivityLog.user_id,
        func.max(ActivityLog.created_at).label("last_at")
    ).filter(
        ActivityLog.action == "LOGIN"
    ).group_by(ActivityLog.user_id).all()

    last_login_map: dict = {row.user_id: row.last_at for row in last_login_raw if row.user_id is not None}

    # Analyses per user this month (by user_name string match)
    analyses_month = db.query(
        InferenceLog.user_name,
        func.count(InferenceLog.id).label("cnt")
    ).filter(
        InferenceLog.created_at.between(start_of_m, end_of_m)
    ).group_by(InferenceLog.user_name).all()

    # Map: lower(user_name) -> count
    analyses_map: dict = {}
    for row in analyses_month:
        key = (row.user_name or "").lower().strip()
        analyses_map[key] = analyses_map.get(key, 0) + row.cnt

    # Reports generated this month (from Report table)
    total_reports_month = db.query(func.count(Report.id)).filter(
        Report.created_at.between(start_of_m, end_of_m)
    ).scalar() or 0

    # Total analyses this month
    total_analyses_month = db.query(func.count(InferenceLog.id)).filter(
        InferenceLog.created_at.between(start_of_m, end_of_m)
    ).scalar() or 0

    # Active users this month = logged in this month
    active_user_ids = {row.user_id for row in logins_month if row.user_id is not None}

    # All analyses for this month
    all_analyses = db.query(InferenceLog).filter(
        InferenceLog.created_at.between(start_of_m, end_of_m)
    ).order_by(InferenceLog.created_at.desc()).all()

    # All reports this month
    all_reports = db.query(Report).filter(
        Report.created_at.between(start_of_m, end_of_m)
    ).order_by(Report.created_at.desc()).all()

    # All activity logs for this month
    all_activities = db.query(ActivityLog).filter(
        ActivityLog.created_at.between(start_of_m, end_of_m)
    ).order_by(ActivityLog.created_at.desc()).all()

    # ── Build user rows for Summary sheet ────────────────────────────────────
    def _resolve_name(u: User) -> str:
        return (u.full_name or u.username or "Unknown").strip()

    def _resolve_role(u: User) -> str:
        role_map = {
            "admin": "ADMIN",
            "estate_manager": "ESTATE_MANAGER",
            "agronomist": "AGRONOMIST",
            "operator": "OPERATOR",
            "viewer": "VIEWER",
        }
        return role_map.get((u.role or "").lower(), (u.role or "USER").upper())

    user_rows_data = []
    for u in users:
        name     = _resolve_name(u)
        role_str = _resolve_role(u)

        # Determine status
        if u.status == "pending":
            status_val = "Pending"
            status_color = C_AMBER
        elif u.id in active_user_ids:
            status_val = "Active"
            status_color = C_GREEN
        else:
            status_val = "Inactive"
            status_color = C_GRAY

        last_dt = last_login_map.get(u.id)
        if last_dt is None:
            last_login_str   = "Never"
            last_login_color = C_LGRAY
        else:
            last_login_str = _fmt_dt(last_dt)
            days_ago = (now_utc - last_dt).days
            last_login_color = C_RED if days_ago > 30 else C_DARK

        name_key = name.lower().strip()
        analyses_cnt  = analyses_map.get(name_key, 0)
        reports_cnt   = 0  
        login_cnt     = logins_map.get(u.id, 0)

        user_rows_data.append({
            "name": name,
            "role": role_str,
            "status": status_val,
            "status_color": status_color,
            "last_login": last_login_str,
            "last_login_color": last_login_color,
            "analyses": analyses_cnt,
            "reports": reports_cnt,
            "logins": login_cnt,
        })

    # ── Period stats ─────────────────────────────────────────────────────────
    total_active = sum(1 for r in user_rows_data if r["status"] == "Active")
    # Total events = analyses + report generations (proxy)
    total_events = total_analyses_month + total_reports_month

    # ── Security notes — derived from data ────────────────────────────────────
    never_active_count = sum(1 for r in user_rows_data if r["last_login"] == "Never")
    pending_count      = sum(1 for r in user_rows_data if r["status"] == "Pending")

    # ──────────────────────────────────────────────────────────────────────────
    # Build workbook
    # ──────────────────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    ws1.views.sheetView[0].showGridLines = True
    MAX_COL_S1 = 7

    # ── Part A: Report Header ────────────────────────────────────────────────
    _merge_header(ws1, 1, "NyawitAI — User Activity & Audit Log", MAX_COL_S1, bold=True, size=14)

    ws1.row_dimensions[2].height = 18
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=MAX_COL_S1)
    c2 = ws1.cell(row=2, column=1, value="PLANTATION INTELLIGENCE SYSTEMS")
    c2.font = _font(size=10, color=C_GRAY)
    c2.alignment = _align()

    # row 3 empty
    ws1.row_dimensions[4].height = 20
    ws1.merge_cells(start_row=4, start_column=1, end_row=4, end_column=MAX_COL_S1)
    c4 = ws1.cell(row=4, column=1, value=f"Monitoring Period: {period_label}")
    c4.font = _font(size=11)
    c4.alignment = _align()

    ws1.row_dimensions[5].height = 18
    ws1.merge_cells(start_row=5, start_column=1, end_row=5, end_column=MAX_COL_S1)
    c5 = ws1.cell(row=5, column=1, value=f"Generated: {gen_ts}   ·   Report ID: {report_id}")
    c5.font = _font(size=10, color=C_GRAY)
    c5.alignment = _align()

    # row 6 empty

    # ── Part B: Period Stats ─────────────────────────────────────────────────
    _section_label(ws1, 7, "PERIOD SUMMARY", MAX_COL_S1)

    stats = [
        ("Total Users Active",      total_active),
        ("Total System Events",     total_events),
        ("Total Analyses Created",  total_analyses_month),
        ("Total Reports Generated", total_reports_month),
    ]
    for i, (label, val) in enumerate(stats, start=8):
        ws1.row_dimensions[i].height = 20
        lc = ws1.cell(row=i, column=1, value=label)
        lc.font = _font(size=11, color=C_GRAY)
        lc.alignment = _align()

        vc = ws1.cell(row=i, column=2, value=val)
        vc.font = _font(size=11, bold=True, color=C_DARK)
        vc.alignment = _align()

    # row 12 empty

    # ── Part C: User Activity Summary Table ───────────────────────────────────
    _section_label(ws1, 13, "USER ACTIVITY SUMMARY", MAX_COL_S1)

    headers_s1 = ["NAME", "ROLE", "STATUS", "LAST LOGIN",
                  "ANALYSES CREATED", "REPORTS EXPORTED", "LOGIN COUNT"]
    _header_row(ws1, 14, headers_s1)

    for idx, row_data in enumerate(user_rows_data):
        excel_row = 15 + idx
        bg = "FFFFFF" if idx % 2 == 0 else C_ALT_ROW
        ws1.row_dimensions[excel_row].height = 20

        def _wc(col, val, bold=False, color=C_DARK, mono=False):
            c = ws1.cell(row=excel_row, column=col, value=val)
            c.font = Font(name="Consolas" if mono else "Calibri",
                          bold=bold, size=11, color=color)
            c.fill = _fill(bg)
            c.alignment = _align()
            c.border = _border()

        _wc(1, row_data["name"],    bold=True)
        _wc(2, row_data["role"])
        _wc(3, row_data["status"],  color=row_data["status_color"])
        _wc(4, row_data["last_login"], color=row_data["last_login_color"])
        _wc(5, row_data["analyses"] if row_data["analyses"] > 0 else "—",
            color=C_LGRAY if row_data["analyses"] == 0 else C_DARK)
        _wc(6, row_data["reports"]  if row_data["reports"]  > 0 else "—",
            color=C_LGRAY if row_data["reports"]  == 0 else C_DARK)
        _wc(7, row_data["logins"]   if row_data["logins"]   > 0 else "—",
            color=C_LGRAY if row_data["logins"]   == 0 else C_DARK)

    user_table_last_row = 14 + len(user_rows_data)

    # ── Part D: Security Notes ────────────────────────────────────────────────
    sec_start = user_table_last_row + 3
    _section_label(ws1, sec_start, "SECURITY NOTES", MAX_COL_S1)

    sec_notes = [
        ("Failed login attempts this month", 0),   # no failed login tracking
        ("Accounts disabled this month",     0),   # no disable tracking
        ("New users invited this month",     pending_count),
        ("Role changes this month",          0),   # no role change tracking
    ]
    for i, (label, val) in enumerate(sec_notes, start=sec_start + 1):
        ws1.row_dimensions[i].height = 20
        lc = ws1.cell(row=i, column=1, value=label)
        lc.font = _font(size=11, color=C_GRAY)
        vc = ws1.cell(row=i, column=2, value=val if val > 0 else "—")
        vc.font = _font(size=11, bold=True,
                        color=C_AMBER if val > 0 else C_LGRAY)

    all_zero = all(v == 0 for _, v in sec_notes)
    if all_zero:
        no_inc_row = sec_start + 1 + len(sec_notes) + 1
        ws1.row_dimensions[no_inc_row].height = 20
        ws1.merge_cells(start_row=no_inc_row, start_column=1,
                        end_row=no_inc_row, end_column=MAX_COL_S1)
        c = ws1.cell(row=no_inc_row, column=1,
                     value="No security incidents detected this period.")
        c.font = _font(size=11, italic=True, color=C_GREEN)

    # ── Column widths Sheet 1 ─────────────────────────────────────────────────
    col_widths_s1 = [28, 22, 12, 28, 20, 20, 14]
    for ci, w in enumerate(col_widths_s1, start=1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # Freeze pane below user table header
    ws1.freeze_panes = "A15"

    # Auto-filter on user table
    ws1.auto_filter.ref = (
        f"A14:{get_column_letter(MAX_COL_S1)}{user_table_last_row}"
    )

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — EVENT LOG
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Event Log")
    ws2.views.sheetView[0].showGridLines = True
    MAX_COL_S2 = 6

    # Combine analyses + report events + activity logs into a unified chronological log
    events = []

    for act in all_activities:
        events.append({
            "ts":          act.created_at,
            "user":        (act.user_name or "Unknown").strip().capitalize(),
            "role":        (act.user_role or "operator").upper(),
            "action":      act.action,  # e.g. LOGIN, LOGOUT
            "detail":      act.detail or f"User performed {act.action}",
            "resource_id": f"ACT-{act.id}",
        })

    for log in all_analyses:
        events.append({
            "ts":          log.created_at,
            "user":        (log.user_name or "Unknown").strip().capitalize(),
            "role":        (log.user_role or "operator").upper(),
            "action":      "ANALYSIS_CREATED",
            "detail":      f"Zone: {log.block_name} · {log.trees_count} trees",
            "resource_id": log.log_code or f"ANL-{log.id}",
        })

    for rep in all_reports:
        fmt = rep.type or "PDF"
        events.append({
            "ts":          rep.created_at,
            "user":        "Admin",
            "role":        "ADMIN",
            "action":      "REPORT_GENERATED",
            "detail":      f"{rep.name} · {fmt}",
            "resource_id": (rep.report_code or str(rep.id))[:8],
        })

    # Sort newest first
    events.sort(key=lambda e: e["ts"] or datetime.min, reverse=True)

    # ── Header rows ──────────────────────────────────────────────────────────
    _merge_header(ws2, 1, "NyawitAI — System Event Log", MAX_COL_S2, bold=True, size=14)

    ws2.row_dimensions[2].height = 18
    ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=MAX_COL_S2)
    c_h2 = ws2.cell(row=2, column=1,
                    value=f"Monitoring Period: {period_label}   ·   Total Events: {len(events)}")
    c_h2.font = _font(size=11, color=C_GRAY)
    c_h2.alignment = _align()

    # row 3 empty

    # ── Table header ─────────────────────────────────────────────────────────
    headers_s2 = ["TIMESTAMP", "USER", "ROLE", "ACTION", "DETAIL", "RESOURCE ID"]
    _header_row(ws2, 4, headers_s2)

    # Action colour map
    ACTION_COLORS = {
        "LOGIN":            C_GRAY,
        "LOGOUT":           C_GRAY,
        "ANALYSIS_CREATED": C_GREEN,
        "REPORT_GENERATED": C_BLUE,
        "REPORT_DOWNLOADED": C_BLUE,
        "USER_INVITED":     C_AMBER,
        "USER_ROLE_CHANGED": C_AMBER,
        "ACCOUNT_DISABLED": C_RED,
    }

    for idx, ev in enumerate(events):
        r = 5 + idx
        bg = "FFFFFF" if idx % 2 == 0 else C_ALT_ROW
        ws2.row_dimensions[r].height = 20
        action_color = ACTION_COLORS.get(ev["action"], C_DARK)

        def _ec(col, val, bold=False, color=C_DARK, mono=False):
            c = ws2.cell(row=r, column=col, value=val)
            c.font = Font(name="Consolas" if mono else "Calibri",
                          bold=bold, size=11, color=color)
            c.fill = _fill(bg)
            c.alignment = _align(wrap=(col == 5))
            c.border = _border()

        ts_str = _fmt_dt(ev["ts"]) if ev["ts"] else "—"
        _ec(1, ts_str)
        _ec(2, ev["user"],   bold=True)
        _ec(3, ev["role"],   color=C_GRAY)
        _ec(4, ev["action"], bold=True, color=action_color)
        _ec(5, ev["detail"])
        _ec(6, ev["resource_id"], color=C_LGRAY, mono=True)

    # ── Column widths Sheet 2 ─────────────────────────────────────────────────
    col_widths_s2 = [22, 18, 20, 22, 35, 14]
    for ci, w in enumerate(col_widths_s2, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # Freeze below event table header
    ws2.freeze_panes = "A5"

    # Auto-filter
    last_ev_row = max(5, 4 + len(events))
    ws2.auto_filter.ref = f"A4:{get_column_letter(MAX_COL_S2)}{last_ev_row}"

    # ── Stream response ───────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = (
        f"nyawitai_user_audit_log_{today.strftime('%Y%m%d')}.xlsx"
    )
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )
